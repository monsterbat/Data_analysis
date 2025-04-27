#!/usr/bin/env python
# -*- coding: utf-8 -*-\
# update : 2025/03/10

import os
import sys
import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


###########################################
#            ColorShift_Analysis          #
###########################################

N = 5  # 這裡可以改成 10，代表要一次處理幾對 U、V

# 建立 2 個列表，分別儲存多筆 U、V 檔案路徑
u_txt_paths = ["" for _ in range(N)]
v_txt_paths = ["" for _ in range(N)]
u_file_labels = []
v_file_labels = []

def choose_u_file(i):
    """選擇第 i 對 U 檔案"""
    global u_txt_paths
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    if file_path:
        u_txt_paths[i] = file_path
        u_file_labels[i].config(text=os.path.basename(file_path))
    else:
        u_txt_paths[i] = ""
        u_file_labels[i].config(text="No file selected")

def choose_v_file(i):
    """選擇第 i 對 V 檔案"""
    global v_txt_paths
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    if file_path:
        v_txt_paths[i] = file_path
        v_file_labels[i].config(text=os.path.basename(file_path))
    else:
        v_txt_paths[i] = ""
        v_file_labels[i].config(text="No file selected")



def load_txt_data(file_path):
    """
    從 TXT 文件中讀取數據，從第 4 行開始解析，並忽略非數字行
    回傳 NumPy 陣列
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()[3:]  # 從第 4 行開始讀取
        data = []
        for line in lines:
            try:
                row = list(map(float, line.split()))
                if row:
                    data.append(row)
            except ValueError:
                continue  # 忽略無法轉換的行
        return np.array(data)
    except Exception as e:
        print(f"讀取檔案時發生錯誤: {e}")
        sys.exit(1)

def modify_data_formula(ws, rows, cols, original_ws):
    """
    利用 openpyxl 直接計算 Modify 工作表的數值，不使用 Excel 公式
    以 Original!A1 為基準，將每個儲存格的值做差
    """
    base_value = original_ws["A1"].value  # 取得 Original!A1 的數值
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            col_letter = get_column_letter(c)
            original_value = original_ws[f"{col_letter}{r}"].value
            if original_value is not None:
                ws[f"{col_letter}{r}"] = original_value - base_value
            else:
                ws[f"{col_letter}{r}"] = None

def calculate_data_formula(ws, modify_ws):
    """
    在 Calculator 工作表中填入數據：
      - A3~A143 由 -70 遞增至 70
      - 設定標題（各欄位不同的參數）
      - 依據 Modify 工作表資料填入其它欄位，包含部分反轉排列與複製
    """
    # 填入 A3 ~ A143
    value = -70
    for r in range(3, 144):
        ws[f"A{r}"] = value
        value += 1

    # 設定標題
    headers = {
        "B": [0, "0~70"], "C": [180, "0~70"], "D": [180, "-70~0"], "E": ["0~180", "FIN"],
        "F": [45, "0~70"], "G": [225, "0~70"], "H": [225, "-70~0"], "I": ["45~225", "FIN"],
        "J": [90, "0~70"], "K": [270, "0~70"], "L": [270, "-70~0"], "M": ["90~270", "FIN"],
        "N": [135, "0~70"], "O": [315, "0~70"], "P": [315, "-70~0"], "Q": ["135~315", "FIN"]
    }
    for col, (h1, h2) in headers.items():
        ws[f"{col}1"] = h1
        ws[f"{col}2"] = h2

    # 填入 B3 ~ B73：根據 Modify 工作表中指定儲存格的數值
    for i in range(71):
        ws[f"B{3+i}"] = modify_ws[f"A{1+i}"].value
        ws[f"C{3+i}"] = modify_ws[f"FY{1+i}"].value
        ws[f"F{3+i}"] = modify_ws[f"AT{1+i}"].value
        ws[f"G{3+i}"] = modify_ws[f"HR{1+i}"].value
        ws[f"J{3+i}"] = modify_ws[f"CM{1+i}"].value
        ws[f"K{3+i}"] = modify_ws[f"JK{1+i}"].value
        ws[f"N{3+i}"] = modify_ws[f"EF{1+i}"].value
        ws[f"O{3+i}"] = modify_ws[f"LD{1+i}"].value

    # 填入 D3 ~ D73 (將 C3~C73 反轉)
    for i in range(71):
        ws[f"D{3+i}"] = ws[f"C{73-i}"].value

    # 填入 H3 ~ H73 (將 G3~G73 反轉)
    for i in range(71):
        ws[f"H{3+i}"] = ws[f"G{73-i}"].value

    # 填入 L3 ~ L73 (將 K3~K73 反轉)
    for i in range(71):
        ws[f"L{3+i}"] = ws[f"K{73-i}"].value

    # 填入 P3 ~ P73 (將 O3~O73 反轉)
    for i in range(71):
        ws[f"P{3+i}"] = ws[f"O{73-i}"].value

    # 填入 E3 ~ E72, I3 ~ I72, M3 ~ M72, Q3 ~ Q72 (複製部分欄位)
    for i in range(70):
        ws[f"E{3+i}"] = ws[f"D{3+i}"].value
        ws[f"I{3+i}"] = ws[f"H{3+i}"].value
        ws[f"M{3+i}"] = ws[f"L{3+i}"].value
        ws[f"Q{3+i}"] = ws[f"P{3+i}"].value

    # 填入 E73 ~ E143 (將 B3~B73 複製)
    for i in range(71):
        ws[f"E{73+i}"] = ws[f"B{3+i}"].value
        ws[f"I{73+i}"] = ws[f"F{3+i}"].value
        ws[f"M{73+i}"] = ws[f"J{3+i}"].value
        ws[f"Q{73+i}"] = ws[f"N{3+i}"].value

def result_data_formula(ws, calculate_ws):
    """
    在 Result 工作表中填入數值，整理 Calculator 工作表的數據：
      - A3~A143 為方向 (-70 到 70)
      - 其餘欄位依據 Calculator 工作表資料填入
    """
    ws["A2"] = "direction"
    value = -70
    for r in range(3, 144):
        ws[f"A{r}"] = value
        value += 1

    ws["B2"] = "0~180"
    ws["C2"] = "45~225"
    ws["D2"] = "90~270"
    ws["E2"] = "135~315"

    for i in range(3, 144):
        ws[f"B{i}"] = calculate_ws[f"E{i}"].value
        ws[f"C{i}"] = calculate_ws[f"I{i}"].value
        ws[f"D{i}"] = calculate_ws[f"M{i}"].value
        ws[f"E{i}"] = calculate_ws[f"Q{i}"].value

def save_to_excel(original, txt_path, uv_label):
    """
    將讀取的原始數據寫入 Excel（含 Original、Modify、Calculate、Result 四個工作表）
    Excel 檔名格式：
      Result_{檔名}_{uv_label}_{日期時間}.xlsx
    回傳儲存後的 Excel 檔案路徑
    """
    original_filename = os.path.splitext(os.path.basename(txt_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d%H%M")

    # 依 uv_label ( 'u' 或 'v' ) 來組合檔名
    output_path = f"{original_filename}_{uv_label}_{timestamp}.xlsx"

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Original"
    for r_idx, row in enumerate(original, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws1.cell(row=r_idx, column=c_idx, value=value)
    
    ws2 = wb.create_sheet("Modify")
    modify_data_formula(ws2, original.shape[0], original.shape[1], ws1)
    
    ws3 = wb.create_sheet("Calculate")
    calculate_data_formula(ws3, ws2)
    
    ws4 = wb.create_sheet("Result")
    result_data_formula(ws4, ws3)

    wb.save(output_path)
    print(f"Excel 檔案已生成，檔名為 {output_path}")
    return output_path

###########################################
#                Combine                  #
###########################################

def read_excel_data(file_path):
    """
    讀取 Excel 檔案中的 'Result' 工作表，取出 B3 到 E143 的數據
    回傳一個 NumPy 陣列
    """
    if not file_path:
        return None
    try:
        xls = pd.ExcelFile(file_path)
        if 'Result' not in xls.sheet_names:
            print(f"錯誤: '{file_path}' 不包含 'Result' 工作表")
            return None
        df = pd.read_excel(file_path, sheet_name='Result', header=None)
        if df.shape[1] < 5:
            print("⚠️ 警告: Excel 檔案的列數少於 5，可能數據沒讀完整！")
        # 取出 B3 到 E143 (注意：pandas 的 index 從 0 開始，所以 row=2~142, col=1~4)
        data = df.iloc[2:143, 1:5].values  
        return data
    except Exception as e:
        print(f"讀取 Excel 檔案時發生錯誤: {e}")
        return None

def calculate_ijkl_values(u_data, v_data):
    """
    利用 U 與 V 數據計算 I、J、K、L 值
    計算公式：sqrt( U^2 + V^2 )
    回傳包含 I, J, K, L 四欄的 DataFrame
    """
    if u_data is None or v_data is None:
        print("❌ 錯誤: 沒有有效的 U 或 V 數據")
        return None
    try:
        u_array = np.array(u_data, dtype=float)
        v_array = np.array(v_data, dtype=float)
    except Exception as e:
        print("❌ 轉換數據時發生錯誤:", e)
        return None

    if u_array.shape[1] < 4 or v_array.shape[1] < 4:
        print("❌ 錯誤: U 或 V 數據的列數不足")
        return None

    I_values = np.sqrt(u_array[:, 0] ** 2 + v_array[:, 0] ** 2)
    J_values = np.sqrt(u_array[:, 1] ** 2 + v_array[:, 1] ** 2)
    K_values = np.sqrt(u_array[:, 2] ** 2 + v_array[:, 2] ** 2)
    L_values = np.sqrt(u_array[:, 3] ** 2 + v_array[:, 3] ** 2)

    ijkl_df = pd.DataFrame({
        'I': I_values,
        'J': J_values,
        'K': K_values,
        'L': L_values
    })
    return ijkl_df

def create_new_excel(u_data, v_data, final_filename="output.xlsx"):
    """
    利用 U 與 V 的數據創建一個新 Excel 檔案，
    並在第 3 行(A3~L3)插入指定文字。
    """
    if u_data is None or v_data is None:
        print("❌ 錯誤: 沒有讀取到有效數據")
        return
    try:
        # 將 u_data, v_data, ijkl_df 分別轉成 DataFrame
        u_df = pd.DataFrame(u_data, columns=['A', 'B', 'C', 'D'])
        v_df = pd.DataFrame(v_data, columns=['E', 'F', 'G', 'H'])

        ijkl_df = calculate_ijkl_values(u_data, v_data)
        if ijkl_df is None:
            print("❌ 錯誤: 無法計算 I、J、K、L 數據")
            return

        # 建立一個只包含一行文字的 DataFrame，用來寫在第 3 行
        # 這裡示範 12 個欄位 (A~L)：
        headers_df = pd.DataFrame([[
            "u 0-180", "u 45-225", "u 90-270", "u 135-315",
            "v 0-180", "v 45-225", "v 90-270", "v 135-315",
            "duv 0-180", "duv 45-225", "duv 90-270", "duv 135-315"
        ]])

        with pd.ExcelWriter(final_filename, engine='openpyxl') as writer:
            # 1) 先寫入那一行文字，指定 startrow=2 => Excel 第 3 行
            headers_df.to_excel(writer, index=False, header=False, startrow=2, startcol=0)

            # 2) 將原本的 U、V、IJKL 三個 DataFrame 改成從第 4 行開始 (startrow=3)
            u_df.to_excel(writer, index=False, header=False, startrow=3, startcol=0)
            v_df.to_excel(writer, index=False, header=False, startrow=3, startcol=4)
            ijkl_df.to_excel(writer, index=False, header=False, startrow=3, startcol=8)

        print(f"✅ 合併結果已儲存至 {final_filename}")
    except Exception as e:
        print("❌ 產生合併 Excel 時發生錯誤:", e)




###########################################
#                主程式                  #
###########################################

def main():
    root = tk.Tk()
    root.title("選取多筆 U 與 V 檔案")

    # 這裡要使用全域的 u_file_labels、v_file_labels
    global u_file_labels, v_file_labels

    for i in range(N):
        # 每一「對」用一個 row_frame 包起來
        row_frame = tk.Frame(root)
        row_frame.pack(fill="x", padx=10, pady=5)

        # 左邊 (U)
        u_frame = tk.LabelFrame(row_frame, text=f"U (第{i+1}組)", padx=10, pady=10)
        u_frame.pack(side="left", fill="x", expand=True)

        # 建立 Label, Button
        u_label = tk.Label(u_frame, text="No file selected")
        u_label.pack(pady=5)
        u_file_labels.append(u_label)  # 存到列表

        u_btn = tk.Button(u_frame, text="Select U File", command=lambda idx=i: choose_u_file(idx))
        u_btn.pack(pady=5)

        # 右邊 (V)
        v_frame = tk.LabelFrame(row_frame, text=f"V (第{i+1}組)", padx=10, pady=10)
        v_frame.pack(side="right", fill="x", expand=True)

        v_label = tk.Label(v_frame, text="No file selected")
        v_label.pack(pady=5)
        v_file_labels.append(v_label)

        v_btn = tk.Button(v_frame, text="Select V File", command=lambda idx=i: choose_v_file(idx))
        v_btn.pack(pady=5)

    # 最底部 Combine 按鈕
    def start_processing():
        # 按下後，逐一檢查每一組 U、V
        for i in range(N):
            u_file = u_txt_paths[i]
            v_file = v_txt_paths[i]
            if u_file and v_file:
                process_files(u_file, v_file)
            else:
                print(f"第{i+1}組未選取完整，跳過...")

        print("所有組別都處理完成。")

    combine_btn = tk.Button(root, text="Combine", command=start_processing)
    combine_btn.pack(side="bottom", pady=10)

    root.mainloop()


def process_files(u_file, v_file):
    analysis_excel_paths = []

    # 第一個檔案 => U
    print("正在處理檔案(U)：", u_file)
    u_original_data = load_txt_data(u_file)
    u_excel_path = save_to_excel(u_original_data, u_file, uv_label="u")
    analysis_excel_paths.append(u_excel_path)

    # 第二個檔案 => V
    print("正在處理檔案(V)：", v_file)
    v_original_data = load_txt_data(v_file)
    v_excel_path = save_to_excel(v_original_data, v_file, uv_label="v")
    analysis_excel_paths.append(v_excel_path)

    print("開始合併分析結果...")
    u_excel = analysis_excel_paths[0]
    v_excel = analysis_excel_paths[1]
    u_data = read_excel_data(u_excel)
    v_data = read_excel_data(v_excel)
    if u_data is None or v_data is None:
        print("無法讀取分析結果進行合併。")
        return

    # 新增：根據 U 檔名，定義最終檔名
    u_basename = os.path.splitext(os.path.basename(u_file))[0]
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    final_filename = f"Result_{u_basename}_{timestamp}.xlsx"

    create_new_excel(u_data, v_data, final_filename)
    print(f"第 {u_basename} 組合併完成，產生檔案：{final_filename}")


if __name__ == "__main__":
    main()